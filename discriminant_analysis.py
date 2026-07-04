import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
from openpyxl import Workbook, load_workbook

from tk_env import configure_tk_environment

configure_tk_environment()

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

SUPPORTED_INPUT_EXTENSIONS = {".csv", ".xlsx"}
LABEL_COLUMN_NAMES = {"cluster", "label", "target"}
RIDGE_RELATIVE_EPSILON = 1e-4
RIDGE_ABSOLUTE_FALLBACK = 1e-9


class StandardScalerModel:
    def __init__(self) -> None:
        self.mean_ = None
        self.var_ = None
        self.scale_ = None

    def fit(self, x_values: np.ndarray) -> "StandardScalerModel":
        self.mean_ = np.mean(x_values, axis=0)
        self.var_ = np.var(x_values, axis=0)
        self.scale_ = np.sqrt(self.var_)
        self.scale_[self.scale_ == 0] = 1.0
        return self

    def transform(self, x_values: np.ndarray) -> np.ndarray:
        if self.mean_ is None or self.scale_ is None:
            raise ValueError("標準化モデルが未学習です。")
        return (x_values - self.mean_) / self.scale_

    def fit_transform(self, x_values: np.ndarray) -> np.ndarray:
        return self.fit(x_values).transform(x_values)


class LinearDiscriminantAnalysisModel:
    def __init__(self) -> None:
        self.classes_ = None
        self.priors_ = None
        self.coef_ = None
        self.intercept_ = None

    def fit(self, x_values: np.ndarray, y_values: np.ndarray) -> "LinearDiscriminantAnalysisModel":
        classes, inverse = np.unique(y_values, return_inverse=True)
        n_samples, n_features = x_values.shape
        n_classes = classes.size

        if n_classes < 2:
            raise ValueError("ラベル列には 2 つ以上のグループが必要です。")

        priors = np.zeros(n_classes, dtype=float)
        means = np.zeros((n_classes, n_features), dtype=float)
        pooled_covariance = np.zeros((n_features, n_features), dtype=float)

        for class_index, class_label in enumerate(classes):
            class_mask = inverse == class_index
            class_values = x_values[class_mask]
            class_count = class_values.shape[0]
            priors[class_index] = class_count / n_samples
            means[class_index] = np.mean(class_values, axis=0)

            centered = class_values - means[class_index]
            pooled_covariance += centered.T @ centered

        denominator = n_samples - n_classes
        if denominator <= 0:
            raise ValueError("学習データ数が不足しているため判別分析を実行できません。")

        pooled_covariance /= denominator

        # 0/1のダミー変数（MA形式）のように相関の強い特徴量が混在すると
        # pooled_covariance がほぼ特異行列になり、固定の極小リッジ(1e-9)では
        # 逆行列が不安定になる。対角成分の平均スケールに対する相対値でリッジを
        # 入れることで、標準化後のスケールに関わらず安定した正則化になる。
        diagonal_scale = float(np.trace(pooled_covariance)) / n_features
        ridge = diagonal_scale * RIDGE_RELATIVE_EPSILON if diagonal_scale > 0 else RIDGE_ABSOLUTE_FALLBACK
        pooled_covariance += np.eye(n_features) * ridge
        covariance_inverse = np.linalg.pinv(pooled_covariance)

        coef = means @ covariance_inverse
        intercept = np.empty(n_classes, dtype=float)
        for class_index in range(n_classes):
            intercept[class_index] = (
                -0.5 * means[class_index] @ covariance_inverse @ means[class_index]
                + np.log(priors[class_index])
            )

        self.classes_ = classes
        self.priors_ = priors
        self.coef_ = coef
        self.intercept_ = intercept
        return self

    def decision_function(self, x_values: np.ndarray) -> np.ndarray:
        if self.coef_ is None or self.intercept_ is None:
            raise ValueError("判別分析モデルが未学習です。")
        return x_values @ self.coef_.T + self.intercept_

    def predict(self, x_values: np.ndarray) -> np.ndarray:
        scores = self.decision_function(x_values)
        best_index = np.argmax(scores, axis=1)
        return self.classes_[best_index]

    def score(self, x_values: np.ndarray, y_values: np.ndarray) -> float:
        predictions = self.predict(x_values)
        return float(np.mean(predictions == y_values))


def leave_one_out_accuracy(x_values: np.ndarray, y_values: np.ndarray):
    """学習データに対する Leave-One-Out 交差検証 (LOOCV) 精度を計算する。

    各サンプルを1件だけ除いた残りのデータで標準化とLDAを再学習し、
    除いたサンプルを予測できるかを検証する。除いた結果クラス数が1つに
    なってしまう（そのクラスの唯一のサンプルだった）場合はそのサンプルを
    評価対象から除外し、evaluated_count に反映する。
    """
    n_samples = x_values.shape[0]
    indices = np.arange(n_samples)
    correct_count = 0
    evaluated_count = 0

    for i in indices:
        mask = indices != i
        y_fold = y_values[mask]
        if np.unique(y_fold).size < 2:
            continue

        scaler = StandardScalerModel()
        x_fold = scaler.fit_transform(x_values[mask])

        lda = LinearDiscriminantAnalysisModel()
        lda.fit(x_fold, y_fold)

        x_left_out = scaler.transform(x_values[i : i + 1])
        prediction = lda.predict(x_left_out)[0]

        evaluated_count += 1
        if prediction == y_values[i]:
            correct_count += 1

    accuracy = correct_count / evaluated_count if evaluated_count else float("nan")
    skipped_count = n_samples - evaluated_count
    return accuracy, evaluated_count, skipped_count


def validate_input_file(file_path: Path) -> Path:
    resolved = Path(file_path).expanduser()
    if not resolved.exists():
        raise FileNotFoundError(f"{resolved} が見つかりません。")
    if resolved.suffix.lower() not in SUPPORTED_INPUT_EXTENSIONS:
        raise ValueError(
            f"{resolved.name} は未対応形式です。対応形式: {', '.join(sorted(SUPPORTED_INPUT_EXTENSIONS))}"
        )
    return resolved


def convert_cell_value(value):
    return "" if value is None else value


def is_header_row(row) -> bool:
    non_empty = [cell for cell in row if str(cell).strip()]
    if len(non_empty) < 2:
        return False
    return isinstance(row[0], str) and bool(str(row[0]).strip())


def normalize_table_rows(rows, file_name: str, header_row_override: Optional[int] = None):
    if header_row_override is not None:
        start_index = header_row_override - 1
        if start_index < 0 or start_index >= len(rows):
            raise ValueError(
                f"{file_name} の見出し行に指定した {header_row_override} 行目が範囲外です"
                f"（データは全 {len(rows)} 行です）。"
            )
    else:
        start_index = None
        for index, row in enumerate(rows):
            if is_header_row(row):
                start_index = index
                break
        if start_index is None:
            raise ValueError(f"{file_name} のヘッダーが読み取れません。")

    normalized_rows = rows[start_index:]
    header = normalized_rows[0]
    data_rows = normalized_rows[1:]
    header_row_number = start_index + 1  # 1行目を1とした、ファイル上の行番号
    return header, data_rows, header_row_number


def load_csv_table(file_path: Path, header_row_override: Optional[int] = None):
    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = [[convert_cell_value(cell) for cell in row] for row in reader]

    if not rows:
        raise ValueError(f"{file_path.name} が空です。")

    header, data_rows, header_row_number = normalize_table_rows(rows, file_path.name, header_row_override)
    return header, data_rows, None, header_row_number


def list_excel_sheet_names(file_path: Path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def load_excel_table(
    file_path: Path,
    sheet_name: Optional[str] = None,
    header_row_override: Optional[int] = None,
):
    workbook = load_workbook(file_path, data_only=True)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"{file_path.name} にシート「{sheet_name}」が見つかりません。"
                f"利用可能なシート: {', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet_name]
    else:
        # workbook.active はExcelで最後に開いていたシートを指すため、
        # 意図せず1シート目以外が読み込まれることがある。
        # シート指定がない場合は常に先頭シート（1シート目）を使う。
        worksheet = workbook.worksheets[0]

    rows = [[convert_cell_value(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]

    if not rows:
        raise ValueError(f"{file_path.name} のシート「{worksheet.title}」が空です。")

    header, data_rows, header_row_number = normalize_table_rows(rows, file_path.name, header_row_override)
    return header, data_rows, worksheet.title, header_row_number


def load_table(
    file_path: Path,
    sheet_name: Optional[str] = None,
    header_row_override: Optional[int] = None,
):
    file_path = validate_input_file(file_path)
    if file_path.suffix.lower() == ".csv":
        return load_csv_table(file_path, header_row_override=header_row_override)
    return load_excel_table(file_path, sheet_name=sheet_name, header_row_override=header_row_override)


def coerce_feature_matrix(
    rows,
    start_col: int,
    end_col: Optional[int] = None,
    feature_names=None,
    file_label: str = "データ",
    header_row_number: int = 1,
) -> np.ndarray:
    sliced_rows = [row[start_col:end_col] for row in rows]
    if not sliced_rows:
        raise ValueError("データ行がありません。")

    try:
        return np.asarray(sliced_rows, dtype=float)
    except (TypeError, ValueError) as error:
        for row_index, row in enumerate(sliced_rows):
            for col_index, value in enumerate(row):
                try:
                    float(value)
                except (TypeError, ValueError):
                    file_row_number = header_row_number + 1 + row_index
                    if feature_names and col_index < len(feature_names):
                        column_label = f"「{feature_names[col_index]}」列"
                    else:
                        column_label = f"{col_index + 1}列目"
                    raise ValueError(
                        f"{file_label} の {file_row_number}行目・{column_label}に数値ではない値 "
                        f"「{value}」が入力されています。特徴量列には数値を入力してください。"
                    ) from error
        # 通常はここに到達しない保険（型変換以外の原因でエラーになった場合）
        raise ValueError(f"{file_label} の特徴量列には数値を入力してください。") from error


def coerce_label_vector(values) -> np.ndarray:
    try:
        return np.asarray(values, dtype=int)
    except ValueError:
        pass

    try:
        return np.asarray(values, dtype=float)
    except ValueError:
        pass

    return np.asarray([str(value) for value in values], dtype=object)


def normalize_header_names(header):
    return [str(name).strip() for name in header]


def split_train_features_label(header, rows, file_label: str = "学習ファイル", header_row_number: int = 1):
    if len(header) < 3:
        raise ValueError("学習データは ID 列、特徴量列、ラベル列の最低 3 列が必要です。")
    if not rows:
        raise ValueError("学習データに行がありません。")

    ids = [row[0] if row else "" for row in rows]
    feature_names = normalize_header_names(header[1:-1])
    x_values = coerce_feature_matrix(
        rows, 1, -1, feature_names=feature_names, file_label=file_label, header_row_number=header_row_number
    )
    y_values = coerce_label_vector([row[-1] for row in rows])
    return ids, x_values, y_values, feature_names


def split_test_features(header, rows, file_label: str = "テストファイル", header_row_number: int = 1):
    if len(header) < 2:
        raise ValueError("テストデータは ID 列と特徴量列の最低 2 列が必要です。")
    if not rows:
        raise ValueError("テストデータに行がありません。")

    ids = [row[0] if row else "" for row in rows]
    last_col_name = str(header[-1]).strip().lower()

    if last_col_name in LABEL_COLUMN_NAMES:
        feature_names = normalize_header_names(header[1:-1])
        x_values = coerce_feature_matrix(
            rows, 1, -1, feature_names=feature_names, file_label=file_label, header_row_number=header_row_number
        )
        y_values = coerce_label_vector([row[-1] for row in rows])
    else:
        feature_names = normalize_header_names(header[1:])
        x_values = coerce_feature_matrix(
            rows, 1, None, feature_names=feature_names, file_label=file_label, header_row_number=header_row_number
        )
        y_values = None

    return ids, x_values, y_values, feature_names


def ensure_matching_features(train_feature_names, test_feature_names):
    if len(train_feature_names) != len(test_feature_names):
        raise ValueError(
            f"学習データの特徴量数 ({len(train_feature_names)}) とテストデータの特徴量数 "
            f"({len(test_feature_names)}) が一致しません。"
        )

    mismatches = [
        f"{index + 1}列目: 学習側=「{train_name}」 / テスト側=「{test_name}」"
        for index, (train_name, test_name) in enumerate(zip(train_feature_names, test_feature_names))
        if train_name != test_name
    ]
    if mismatches:
        raise ValueError(
            "学習データとテストデータで特徴量の項目名（並び順）が一致しません。\n"
            + "\n".join(mismatches)
        )


def ensure_valid(y_values):
    unique = np.unique(y_values)
    if unique.size < 2:
        raise ValueError("ラベル列には 2 つ以上のグループが必要です。")
    return unique


def timestamp_suffix() -> str:
    return datetime.now().strftime("%m%d%H%M")


def add_timestamp_to_path(base_path: Path, stamp: Optional[str] = None) -> Path:
    stamp = stamp or timestamp_suffix()
    parent = base_path.parent if base_path.parent != Path("") else Path.cwd()
    return parent / f"{base_path.stem}_{stamp}{base_path.suffix}"


def ensure_unique_path(path: Path) -> Path:
    if not path.exists():
        return path

    counter = 1
    while True:
        candidate = path.with_name(f"{path.stem}_{counter:02d}{path.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def default_output_paths(test_path: Path, output_dir: Optional[Path] = None, stamp: Optional[str] = None):
    stamp = stamp or timestamp_suffix()
    output_dir = Path(output_dir) if output_dir else test_path.parent
    test_extension = test_path.suffix.lower() if test_path.suffix.lower() in SUPPORTED_INPUT_EXTENSIONS else ".csv"
    report_path = ensure_unique_path(add_timestamp_to_path(output_dir / f"lda_report{test_extension}", stamp))
    test_output_path = ensure_unique_path(
        add_timestamp_to_path(output_dir / f"{test_path.stem}_with_prediction{test_extension}", stamp)
    )
    return report_path, test_output_path


def write_csv_table(output_path: Path, header, rows) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerows(rows)


def write_excel_table(output_path: Path, header, rows) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(list(header))
    for row in rows:
        worksheet.append(list(row))
    workbook.save(output_path)


def save_table(output_path: Path, header, rows) -> None:
    suffix = output_path.suffix.lower()
    if suffix == ".csv":
        write_csv_table(output_path, header, rows)
        return
    if suffix == ".xlsx":
        write_excel_table(output_path, header, rows)
        return
    raise ValueError(f"{output_path.name} の出力形式は未対応です。")


def _fmt(value):
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, np.integer):
        return int(value)
    return value


def build_crosstab_rows(pad, true_values: np.ndarray, predicted_values: np.ndarray, predicted_classes: np.ndarray):
    rows = []
    true_classes = np.unique(true_values)
    rows.append(pad(["クラスター", *list(predicted_classes), "合計"]))
    column_totals = np.zeros(len(predicted_classes), dtype=int)
    for true_label in true_classes:
        counts = []
        true_mask = true_values == true_label
        for pred_index, predicted_label in enumerate(predicted_classes):
            count = int(np.sum(true_mask & (predicted_values == predicted_label)))
            counts.append(count)
            column_totals[pred_index] += count
        rows.append(pad([true_label, *counts, int(np.sum(counts))]))
    rows.append(pad(["合計", *[int(total) for total in column_totals], int(np.sum(column_totals))]))
    return rows


def compute_class_variances(x_values: np.ndarray, y_values: np.ndarray, classes: np.ndarray) -> np.ndarray:
    n_features = x_values.shape[1]
    variances = np.zeros((len(classes), n_features), dtype=float)
    for class_index, class_label in enumerate(classes):
        variances[class_index] = np.var(x_values[y_values == class_label], axis=0)
    return variances


def build_class_variance_comment(classes: np.ndarray, avg_variance: np.ndarray) -> str:
    max_index = int(np.argmax(avg_variance))
    min_index = int(np.argmin(avg_variance))
    max_class, min_class = classes[max_index], classes[min_index]
    max_value, min_value = float(avg_variance[max_index]), float(avg_variance[min_index])

    if min_value <= 0:
        return (
            f"クラス「{min_class}」の平均分散がほぼ0のため、比率による比較ができません。"
            "サンプル数が極端に少ない、または値がほぼ一定になっている可能性があります。"
        )

    ratio = max_value / min_value
    if ratio >= 3:
        judgement = (
            "分散に非常に大きな差があります。LDAが仮定する「クラス間で分散が等しい」という前提から"
            "外れている可能性が高く、分散が大きいクラスと小さいクラスの境界で誤判別が増えやすいので注意してください。"
        )
    elif ratio >= 1.5:
        judgement = (
            "分散にやや差があります。大きな問題にはなりにくいですが、境界付近のクラスでは誤判別が"
            "増える可能性があります。"
        )
    else:
        judgement = "クラス間の分散に大きな差は見られません。LDAの前提に対して大きな問題はないと考えられます。"

    return (
        f"{judgement}"
        f"（最大: クラス「{max_class}」平均分散 {max_value:.4f} / "
        f"最小: クラス「{min_class}」平均分散 {min_value:.4f} / 比率 {ratio:.2f}倍）"
    )


def build_report_rows(
    train_accuracy: float,
    lda: LinearDiscriminantAnalysisModel,
    scaler: StandardScalerModel,
    y_train: np.ndarray,
    train_predictions: np.ndarray,
    test_predictions: np.ndarray,
    loocv_accuracy: float,
    loocv_evaluated_count: int,
    loocv_skipped_count: int,
    x_train_std: np.ndarray,
    y_test: Optional[np.ndarray] = None,
    test_accuracy: Optional[float] = None,
):
    n_features = lda.coef_.shape[1]
    n_cols = max(n_features + 2, len(lda.classes_) + 2, 4)

    def pad(row):
        row = list(row)
        return row + [""] * (n_cols - len(row))

    empty = pad([])
    header = pad(["item", "value"])

    rows = []
    rows.append(pad(["学習判別率（再代入精度）", _fmt(train_accuracy)]))
    rows.append(
        pad(
            [
                "注記",
                "学習判別率は学習データ自体への当てはめ精度であり、新規データでの精度を保証するものではありません。",
            ]
        )
    )
    rows.append(pad(["Leave-One-Out交差検証精度(LOOCV)", _fmt(loocv_accuracy)]))
    if loocv_skipped_count > 0:
        rows.append(
            pad(
                [
                    "(LOOCV評価除外件数)",
                    loocv_skipped_count,
                    f"評価対象 {loocv_evaluated_count} 件 / 除外はクラスの唯一のサンプルだったため",
                ]
            )
        )
    rows.append(empty)
    rows.append(pad(["学習データクロス集計", "予測グループ"]))
    rows.extend(build_crosstab_rows(pad, y_train, train_predictions, lda.classes_))
    rows.append(empty)
    rows.append(pad(["標準化", "平均=0, 分散=1 (自前 StandardScaler 適用済み)"]))
    rows.append(empty)
    rows.append(pad(["クラスラベル", *list(lda.classes_)]))
    rows.append(pad(["クラス事前確率", *[_fmt(p) for p in lda.priors_]]))
    rows.append(empty)
    rows.append(["クラス", *[f"coef_{index + 1}" for index in range(n_features)]])
    for index, class_label in enumerate(lda.classes_):
        rows.append([f"クラス {class_label}", *[_fmt(c) for c in lda.coef_[index]]])
    rows.append(empty)
    rows.append(pad(["クラス", "intercept"]))
    for index, class_label in enumerate(lda.classes_):
        rows.append(pad([f"クラス {class_label}", _fmt(lda.intercept_[index])]))
    rows.append(empty)
    rows.append(["標準化平均", *[f"mean_{index + 1}" for index in range(n_features)]])
    rows.append(["値", *[_fmt(m) for m in scaler.mean_]])
    rows.append(empty)
    rows.append(["標準化分散", *[f"var_{index + 1}" for index in range(n_features)]])
    rows.append(["値", *[_fmt(v) for v in scaler.var_]])
    rows.append(empty)

    class_variances = compute_class_variances(x_train_std, y_train, lda.classes_)
    class_avg_variance = class_variances.mean(axis=1)
    rows.append(
        pad(
            [
                "クラス別分散（標準化後）",
                "LDAは全クラスで分散が等しいと仮定するため、クラス間で大きく異なる場合は誤判別が増えやすくなります",
            ]
        )
    )
    rows.append(["クラス", *[f"var_{index + 1}" for index in range(n_features)], "平均分散"])
    for class_index, class_label in enumerate(lda.classes_):
        rows.append(
            [
                f"クラス {class_label}",
                *[_fmt(v) for v in class_variances[class_index]],
                _fmt(class_avg_variance[class_index]),
            ]
        )
    rows.append(pad(["コメント", build_class_variance_comment(lda.classes_, class_avg_variance)]))
    rows.append(empty)
    rows.append(pad(["テストデータ予測グループ集計分布"]))
    rows.append(pad(["predicted_label", "count", "ratio"]))
    test_total = len(test_predictions)
    for class_label in lda.classes_:
        count = int(np.sum(test_predictions == class_label))
        ratio = count / test_total if test_total else 0
        rows.append(pad([class_label, count, _fmt(ratio)]))
    rows.append(pad(["total", test_total, _fmt(1.0 if test_total else 0)]))

    if test_accuracy is not None:
        rows.append(empty)
        rows.append(
            pad(
                [
                    "テスト判別率",
                    _fmt(test_accuracy),
                    "テストファイル末尾の cluster/label/target 列を正解として検証",
                ]
            )
        )
        rows.append(pad(["テストデータクロス集計", "予測グループ"]))
        rows.extend(build_crosstab_rows(pad, y_test, test_predictions, lda.classes_))

    return header, rows


def append_prediction_column(header, rows, predictions):
    output_header = list(header) + ["predicted_label"]
    output_rows = [list(row) + [prediction] for row, prediction in zip(rows, predictions)]
    return output_header, output_rows


def run_analysis(
    train_path: Path,
    test_path: Path,
    output_dir: Optional[Path] = None,
    output_report: Optional[Path] = None,
    output_test: Optional[Path] = None,
    train_sheet: Optional[str] = None,
    test_sheet: Optional[str] = None,
    train_header_row: Optional[int] = None,
    test_header_row: Optional[int] = None,
):
    train_header, train_rows, train_sheet_used, train_header_row = load_table(
        train_path, sheet_name=train_sheet, header_row_override=train_header_row
    )
    test_header, test_rows, test_sheet_used, test_header_row = load_table(
        test_path, sheet_name=test_sheet, header_row_override=test_header_row
    )

    _, x_train, y_train, train_feature_names = split_train_features_label(
        train_header, train_rows, file_label=train_path.name, header_row_number=train_header_row
    )
    _, x_test, y_test, test_feature_names = split_test_features(
        test_header, test_rows, file_label=test_path.name, header_row_number=test_header_row
    )

    ensure_valid(y_train)
    ensure_matching_features(train_feature_names, test_feature_names)

    if output_dir:
        output_dir = Path(output_dir).expanduser()
        output_dir.mkdir(parents=True, exist_ok=True)

    stamp = timestamp_suffix()
    default_report_path, default_test_output_path = default_output_paths(Path(test_path), output_dir, stamp=stamp)
    report_path = (
        ensure_unique_path(add_timestamp_to_path(Path(output_report), stamp)) if output_report else default_report_path
    )
    test_output_path = (
        ensure_unique_path(add_timestamp_to_path(Path(output_test), stamp)) if output_test else default_test_output_path
    )
    report_path.parent.mkdir(parents=True, exist_ok=True)
    test_output_path.parent.mkdir(parents=True, exist_ok=True)

    scaler = StandardScalerModel()
    x_train_std = scaler.fit_transform(x_train)
    x_test_std = scaler.transform(x_test)

    lda = LinearDiscriminantAnalysisModel()
    lda.fit(x_train_std, y_train)

    y_pred = lda.predict(x_test_std)
    train_predictions = lda.predict(x_train_std)
    train_accuracy = float(np.mean(train_predictions == y_train))
    loocv_accuracy, loocv_evaluated_count, loocv_skipped_count = leave_one_out_accuracy(
        x_train, y_train
    )

    test_accuracy = float(np.mean(y_pred == y_test)) if y_test is not None else None

    report_header, report_rows = build_report_rows(
        train_accuracy,
        lda,
        scaler,
        y_train,
        train_predictions,
        y_pred,
        loocv_accuracy,
        loocv_evaluated_count,
        loocv_skipped_count,
        x_train_std,
        y_test,
        test_accuracy,
    )
    save_table(report_path, report_header, report_rows)

    test_output_header, test_output_rows = append_prediction_column(test_header, test_rows, y_pred)
    save_table(test_output_path, test_output_header, test_output_rows)

    return {
        "train_accuracy": train_accuracy,
        "loocv_accuracy": loocv_accuracy,
        "loocv_evaluated_count": loocv_evaluated_count,
        "loocv_skipped_count": loocv_skipped_count,
        "test_accuracy": test_accuracy,
        "train_sheet_used": train_sheet_used,
        "test_sheet_used": test_sheet_used,
        "train_header_row": train_header_row,
        "test_header_row": test_header_row,
        "report_path": report_path,
        "test_output_path": test_output_path,
    }


class DiscriminantAnalysisApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("判別分析 (LDA)")
        self.root.geometry("1040x300")
        self.root.minsize(1000, 280)

        self.train_var = tk.StringVar()
        self.test_var = tk.StringVar()
        self.output_dir_var = tk.StringVar()
        self.train_sheet_var = tk.StringVar()
        self.test_sheet_var = tk.StringVar()
        self.train_header_row_var = tk.StringVar()
        self.test_header_row_var = tk.StringVar()
        self.status_var = tk.StringVar(value="学習用ファイルとテスト用ファイルを選択してください。")

        self._build_widgets()

    def _build_widgets(self) -> None:
        frame = ttk.Frame(self.root, padding=16)
        frame.pack(fill="both", expand=True)

        self.root.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="学習ファイル").grid(row=0, column=0, sticky="w", pady=(0, 8))
        ttk.Entry(frame, textvariable=self.train_var).grid(row=0, column=1, sticky="ew", padx=8, pady=(0, 8))
        ttk.Button(frame, text="参照", command=self.select_train_file).grid(row=0, column=2, pady=(0, 8))
        ttk.Label(frame, text="シート").grid(row=0, column=3, sticky="w", padx=(12, 4), pady=(0, 8))
        self.train_sheet_combo = ttk.Combobox(
            frame, textvariable=self.train_sheet_var, width=16, state="disabled"
        )
        self.train_sheet_combo.grid(row=0, column=4, pady=(0, 8))
        ttk.Label(frame, text="見出し行").grid(row=0, column=5, sticky="w", padx=(12, 4), pady=(0, 8))
        ttk.Entry(frame, textvariable=self.train_header_row_var, width=6).grid(row=0, column=6, pady=(0, 8))

        ttk.Label(frame, text="テストファイル").grid(row=1, column=0, sticky="w", pady=(0, 8))
        ttk.Entry(frame, textvariable=self.test_var).grid(row=1, column=1, sticky="ew", padx=8, pady=(0, 8))
        ttk.Button(frame, text="参照", command=self.select_test_file).grid(row=1, column=2, pady=(0, 8))
        ttk.Label(frame, text="シート").grid(row=1, column=3, sticky="w", padx=(12, 4), pady=(0, 8))
        self.test_sheet_combo = ttk.Combobox(
            frame, textvariable=self.test_sheet_var, width=16, state="disabled"
        )
        self.test_sheet_combo.grid(row=1, column=4, pady=(0, 8))
        ttk.Label(frame, text="見出し行").grid(row=1, column=5, sticky="w", padx=(12, 4), pady=(0, 8))
        ttk.Entry(frame, textvariable=self.test_header_row_var, width=6).grid(row=1, column=6, pady=(0, 8))

        ttk.Label(frame, text="出力フォルダ").grid(row=2, column=0, sticky="w", pady=(0, 8))
        ttk.Entry(frame, textvariable=self.output_dir_var).grid(
            row=2, column=1, columnspan=3, sticky="ew", padx=8, pady=(0, 8)
        )
        ttk.Button(frame, text="参照", command=self.select_output_dir).grid(row=2, column=4, pady=(0, 8))

        note = (
            "対応形式: csv / xlsx\n"
            "xlsxで複数シートがある場合は「シート」で読み込む対象を選択してください（未指定時は先頭シートを使用します）。\n"
            "「見出し行」は項目名が入っている行番号です（未指定時は自動検出）。自動検出が誤っている場合のみ入力してください。\n"
            "出力ファイル名には月日時分のタイムスタンプを自動付与し、既存ファイルを上書きしません。"
        )
        ttk.Label(frame, text=note, justify="left").grid(row=3, column=0, columnspan=7, sticky="w", pady=(6, 16))

        button_frame = ttk.Frame(frame)
        button_frame.grid(row=4, column=0, columnspan=7, sticky="ew")
        button_frame.columnconfigure(0, weight=1)

        self.run_button = ttk.Button(button_frame, text="実行", command=self.run)
        self.run_button.grid(row=0, column=0, sticky="e")

        ttk.Separator(frame).grid(row=5, column=0, columnspan=7, sticky="ew", pady=12)
        ttk.Label(frame, textvariable=self.status_var, justify="left").grid(row=6, column=0, columnspan=7, sticky="w")

    def select_train_file(self) -> None:
        path = self._ask_open_file()
        if path:
            self.train_var.set(path)
            self._fill_output_dir_from_input(path)
            self._update_sheet_options(path, self.train_sheet_var, self.train_sheet_combo)

    def select_test_file(self) -> None:
        path = self._ask_open_file()
        if path:
            self.test_var.set(path)
            self._fill_output_dir_from_input(path)
            self._update_sheet_options(path, self.test_sheet_var, self.test_sheet_combo)

    def select_output_dir(self) -> None:
        selected = filedialog.askdirectory(
            title="出力フォルダを選択",
            initialdir=self.output_dir_var.get() or str(Path.cwd()),
        )
        if selected:
            self.output_dir_var.set(selected)

    def _fill_output_dir_from_input(self, selected_path: str) -> None:
        if not self.output_dir_var.get():
            self.output_dir_var.set(str(Path(selected_path).parent))

    def _update_sheet_options(
        self, selected_path: str, sheet_var: tk.StringVar, sheet_combo: ttk.Combobox
    ) -> None:
        sheet_var.set("")
        if Path(selected_path).suffix.lower() != ".xlsx":
            sheet_combo.configure(values=[], state="disabled")
            return

        try:
            sheet_names = list_excel_sheet_names(Path(selected_path))
        except Exception as error:
            sheet_combo.configure(values=[], state="disabled")
            messagebox.showerror(
                "シート取得エラー", f"シート一覧を取得できませんでした。\n{error}"
            )
            return

        sheet_combo.configure(values=sheet_names, state="readonly" if sheet_names else "disabled")
        if sheet_names:
            sheet_var.set(sheet_names[0])

    def _ask_open_file(self) -> str:
        filetypes = [
            ("対応ファイル", "*.csv *.xlsx"),
            ("CSV ファイル", "*.csv"),
            ("Excel ファイル", "*.xlsx"),
        ]
        return filedialog.askopenfilename(
            title="ファイルを選択",
            filetypes=filetypes,
            initialdir=self.output_dir_var.get() or str(Path.cwd()),
        )

    @staticmethod
    def _parse_header_row(raw_value: str, field_label: str) -> Optional[int]:
        raw_value = raw_value.strip()
        if not raw_value:
            return None
        try:
            value = int(raw_value)
        except ValueError as error:
            raise ValueError(f"{field_label}には数値（行番号）を入力してください。") from error
        if value < 1:
            raise ValueError(f"{field_label}には 1 以上の値を入力してください。")
        return value

    def run(self) -> None:
        train_path = self.train_var.get().strip()
        test_path = self.test_var.get().strip()
        output_dir = self.output_dir_var.get().strip()

        if not train_path or not test_path:
            messagebox.showerror("入力不足", "学習ファイルとテストファイルを選択してください。")
            return

        try:
            train_header_row = self._parse_header_row(self.train_header_row_var.get(), "学習ファイルの見出し行")
            test_header_row = self._parse_header_row(self.test_header_row_var.get(), "テストファイルの見出し行")
        except ValueError as error:
            messagebox.showerror("入力エラー", str(error))
            return

        self.run_button.state(["disabled"])
        self.status_var.set("処理を実行しています...")
        self.root.update_idletasks()

        try:
            result = run_analysis(
                train_path=Path(train_path),
                test_path=Path(test_path),
                output_dir=Path(output_dir) if output_dir else None,
                train_sheet=self.train_sheet_var.get().strip() or None,
                test_sheet=self.test_sheet_var.get().strip() or None,
                train_header_row=train_header_row,
                test_header_row=test_header_row,
            )
        except FileNotFoundError as error:
            messagebox.showerror("ファイルエラー", str(error))
            self.status_var.set(str(error))
        except ValueError as error:
            messagebox.showerror("入力エラー", str(error))
            self.status_var.set(str(error))
        except Exception as error:
            messagebox.showerror("予期せぬエラー", str(error))
            self.status_var.set(str(error))
        else:
            summary = ""
            if result["train_sheet_used"]:
                summary += f"学習ファイルのシート: {result['train_sheet_used']}\n"
            if result["test_sheet_used"]:
                summary += f"テストファイルのシート: {result['test_sheet_used']}\n"
            summary += (
                f"学習ファイルの見出し行: {result['train_header_row']}行目\n"
                f"テストファイルの見出し行: {result['test_header_row']}行目\n"
            )
            summary += (
                f"学習判別率（再代入精度）: {result['train_accuracy']:.6f}\n"
                f"LOOCV交差検証精度: {result['loocv_accuracy']:.6f}\n"
            )
            if result["test_accuracy"] is not None:
                summary += f"テスト判別率: {result['test_accuracy']:.6f}\n"
            summary += (
                f"レポート: {result['report_path']}\n"
                f"予測結果: {result['test_output_path']}"
            )
            messagebox.showinfo("完了", summary)
            self.status_var.set(summary)
        finally:
            self.run_button.state(["!disabled"])


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="判別分析ツール (LDA)")
    parser.add_argument("--train", type=Path, help="学習用ファイルパス")
    parser.add_argument("--test", type=Path, help="テスト用ファイルパス")
    parser.add_argument("--output-dir", type=Path, help="出力先フォルダ")
    parser.add_argument("--output-report", type=Path, help="レポート出力ファイル")
    parser.add_argument("--output-test", type=Path, help="予測結果出力ファイル")
    parser.add_argument("--train-sheet", type=str, help="学習ファイル(xlsx)のシート名（省略時は先頭シート）")
    parser.add_argument("--test-sheet", type=str, help="テストファイル(xlsx)のシート名（省略時は先頭シート）")
    parser.add_argument(
        "--train-header-row", type=int, help="学習ファイルの見出し行番号（省略時は自動検出）"
    )
    parser.add_argument(
        "--test-header-row", type=int, help="テストファイルの見出し行番号（省略時は自動検出）"
    )
    parser.add_argument("--gui", action="store_true", help="GUI を起動する")
    return parser


def launch_gui() -> None:
    root = tk.Tk()
    style = ttk.Style()
    if "vista" in style.theme_names():
        style.theme_use("vista")
    DiscriminantAnalysisApp(root)
    root.mainloop()


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if args.gui or len(sys.argv) == 1:
        launch_gui()
        return

    if not args.train or not args.test:
        parser.error("--train と --test を指定するか、引数なしで GUI を起動してください。")

    try:
        result = run_analysis(
            train_path=args.train,
            test_path=args.test,
            output_dir=args.output_dir,
            output_report=args.output_report,
            output_test=args.output_test,
            train_sheet=args.train_sheet,
            test_sheet=args.test_sheet,
            train_header_row=args.train_header_row,
            test_header_row=args.test_header_row,
        )
        print("===== 判別分析完了 =====")
        if result["train_sheet_used"]:
            print(f"学習ファイルのシート: {result['train_sheet_used']}")
        if result["test_sheet_used"]:
            print(f"テストファイルのシート: {result['test_sheet_used']}")
        print(f"学習ファイルの見出し行: {result['train_header_row']}行目")
        print(f"テストファイルの見出し行: {result['test_header_row']}行目")
        print(f"学習判別率（再代入精度）: {result['train_accuracy']:.6f}")
        print(f"LOOCV交差検証精度: {result['loocv_accuracy']:.6f}")
        if result["test_accuracy"] is not None:
            print(f"テスト判別率: {result['test_accuracy']:.6f}")
        print(f"レポート出力: {result['report_path']}")
        print(f"テスト出力: {result['test_output_path']}")
    except FileNotFoundError as error:
        print(f"ファイルが見つかりません: {error}", file=sys.stderr)
        sys.exit(1)
    except ValueError as error:
        print(f"入力データ検証エラー: {error}", file=sys.stderr)
        sys.exit(1)
    except Exception as error:
        print(f"予期せぬエラー: {error}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
